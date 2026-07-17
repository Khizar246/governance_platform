"""Excel-export *content* tests, focused on the formula-injection guard.

Sacred Rule: every export keeps strings_to_formulas=False (and
strings_to_urls=False), so a cell whose value starts with '=' is written as
literal text, never a live formula. Two distinct writer configs enforce this:

  * pandas ExcelWriter(engine_kwargs=...) — Oracle Comparator, Ruleset Mapping
  * a direct xlsxwriter.Workbook(options) — SOD & SA export_results

Both are exercised end-to-end here by feeding a '='-leading payload through the
real export and reading the workbook back with openpyxl: the cell must come
back as a string (data_type 's'), not a formula ('f'). The SOD & SA case also
pins the produced sheet name, which had no content coverage before.
"""

import io

import openpyxl
import polars as pl

from engines.oracle_comparator_engine import run_analysis, generate_report
from engines.sod_sa_engine import export_results, ROLE_OUTPUT_COLUMNS

_PAYLOAD = "=cmd|'/c calc'!A1"


def _leading_equals_cells(workbook_bytes: bytes):
    """Return (sheet, coord, data_type) for every cell whose value starts with '='."""
    wb = openpyxl.load_workbook(io.BytesIO(workbook_bytes))
    return [
        (ws.title, cell.coordinate, cell.data_type)
        for ws in wb.worksheets
        for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]


# ── Oracle Comparator: pandas ExcelWriter guard ───────────────────────────────

def test_oracle_export_writes_formula_payload_as_text():
    env1 = pl.DataFrame({
        "ROLE NAME":           [_PAYLOAD, "GL User"],
        "ENTITLEMENT":         ["=1+1", "Post Journal"],
        "INHERITED ROLE NAME": ["AP Clerk", "GL Base"],
    })
    env2 = pl.DataFrame({
        "ROLE NAME":           ["AP Manager"],
        "ENTITLEMENT":         ["Create Invoice"],
        "INHERITED ROLE NAME": ["AP Clerk"],
    })
    res = run_analysis({"rbac_file1": env1, "rbac_file2": env2}, "rbac", "Prod", "UAT")
    report = generate_report(res.data[0], res.data[1], "Prod", "UAT")
    assert report.success

    cells = _leading_equals_cells(report.data)
    assert cells, "expected the '='-leading payload to survive into the workbook"
    # Every one must be a literal string, never a formula.
    assert all(dtype == "s" for _, _, dtype in cells)


# ── SOD & SA: direct xlsxwriter.Workbook guard ────────────────────────────────

def _role_sod_frame() -> pl.DataFrame:
    # One row; put the injection payload in the first text column.
    return pl.DataFrame({
        col: [_PAYLOAD if i == 0 else "x"]
        for i, col in enumerate(ROLE_OUTPUT_COLUMNS)
    })


def test_sod_sa_export_writes_formula_payload_as_text():
    empty = pl.DataFrame()
    res = export_results(
        role_sod_violations=_role_sod_frame(),
        role_sa_violations=empty,
        user_sod_violations=empty,
        user_sa_violations=empty,
        sod_controls_df=empty,
        sa_controls_df=empty,
        analysis_type="role",
    )
    assert res.success

    data = res.data.getvalue()
    cells = _leading_equals_cells(data)
    assert cells, "expected the '='-leading payload to survive into the workbook"
    assert all(dtype == "s" for _, _, dtype in cells)


def _role_sod_header(fp_enabled: bool):
    empty = pl.DataFrame()
    res = export_results(
        role_sod_violations=_role_sod_frame(),
        role_sa_violations=empty,
        user_sod_violations=empty,
        user_sa_violations=empty,
        sod_controls_df=empty,
        sa_controls_df=empty,
        analysis_type="role",
        fp_enabled=fp_enabled,
    )
    wb = openpyxl.load_workbook(res.data)
    assert "ROLE_SOD" in wb.sheetnames
    return [c.value for c in next(wb["ROLE_SOD"].iter_rows(max_row=1))]


def test_sod_sa_role_header_fp_enabled_is_full_column_order():
    # With FP on, the header is the exact ROLE_OUTPUT_COLUMNS order (Sacred Rule).
    assert _role_sod_header(fp_enabled=True) == ROLE_OUTPUT_COLUMNS


def test_sod_sa_role_header_no_fp_omits_fp_columns():
    # Without FP, the two trailing FP columns are dropped; the rest keep order.
    assert _role_sod_header(fp_enabled=False) == ROLE_OUTPUT_COLUMNS[:-2]
    assert ROLE_OUTPUT_COLUMNS[-2:] == ["Potential FP", "Reason"]
