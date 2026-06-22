#!/usr/bin/env python3
"""Generate expanded SOD & SA Analysis datasets with proper entitlement-to-privilege mapping and real conflicts."""

import random
from pathlib import Path
from openpyxl import Workbook

# Seed for reproducibility
random.seed(42)

# Define entitlements (business capabilities)
ENTITLEMENTS = [
    'Create Purchase Orders',
    'Approve Purchase Orders',
    'Create Vendor Master',
    'Modify Vendor Master',
    'Create Invoices',
    'Approve Invoices',
    'Post Journal Entries',
    'Approve Journal Entries',
    'Close General Ledger',
    'Create Employee Records',
    'Modify Employee Records',
    'Process Payroll',
    'Approve Payroll',
    'Access Bank Accounts',
    'Reconcile Bank Accounts',
    'Create Fixed Assets',
    'Depreciate Fixed Assets',
    'Create Accounts Receivable',
    'Collect Accounts Receivable',
    'System Administration',
    'User Access Management',
    'Audit Log Access',
    'Modify Audit Logs',
    'Generate Compliance Reports',
    'Authorize Transactions',
    'View Sensitive Data',
    'Export Financial Data',
    'Create Purchase Requisitions',
    'Approve Purchase Requisitions',
    'Release Purchase Orders'
]

# Define privilege codes that map to entitlements
# Each entitlement maps to 2-4 privileges
ENTITLEMENT_PRIVILEGE_MAP = {
    'Create Purchase Orders': ['PO_CREATE_01', 'PO_CREATE_02', 'PO_CREATE_03'],
    'Approve Purchase Orders': ['PO_APPROVE_01', 'PO_APPROVE_02', 'PO_APPROVE_03'],
    'Create Vendor Master': ['VENDOR_CREATE_01', 'VENDOR_CREATE_02'],
    'Modify Vendor Master': ['VENDOR_MODIFY_01', 'VENDOR_MODIFY_02', 'VENDOR_MODIFY_03'],
    'Create Invoices': ['INV_CREATE_01', 'INV_CREATE_02', 'INV_CREATE_03'],
    'Approve Invoices': ['INV_APPROVE_01', 'INV_APPROVE_02', 'INV_APPROVE_03'],
    'Post Journal Entries': ['JE_POST_01', 'JE_POST_02', 'JE_POST_03'],
    'Approve Journal Entries': ['JE_APPROVE_01', 'JE_APPROVE_02', 'JE_APPROVE_03'],
    'Close General Ledger': ['GL_CLOSE_01', 'GL_CLOSE_02', 'GL_CLOSE_03'],
    'Create Employee Records': ['EMP_CREATE_01', 'EMP_CREATE_02', 'EMP_CREATE_03'],
    'Modify Employee Records': ['EMP_MODIFY_01', 'EMP_MODIFY_02', 'EMP_MODIFY_03'],
    'Process Payroll': ['PAY_PROCESS_01', 'PAY_PROCESS_02', 'PAY_PROCESS_03'],
    'Approve Payroll': ['PAY_APPROVE_01', 'PAY_APPROVE_02', 'PAY_APPROVE_03'],
    'Access Bank Accounts': ['BANK_ACCESS_01', 'BANK_ACCESS_02', 'BANK_ACCESS_03'],
    'Reconcile Bank Accounts': ['BANK_RECON_01', 'BANK_RECON_02', 'BANK_RECON_03'],
    'Create Fixed Assets': ['FA_CREATE_01', 'FA_CREATE_02', 'FA_CREATE_03'],
    'Depreciate Fixed Assets': ['FA_DEPRECIATE_01', 'FA_DEPRECIATE_02', 'FA_DEPRECIATE_03'],
    'Create Accounts Receivable': ['AR_CREATE_01', 'AR_CREATE_02', 'AR_CREATE_03'],
    'Collect Accounts Receivable': ['AR_COLLECT_01', 'AR_COLLECT_02', 'AR_COLLECT_03'],
    'System Administration': ['SYSADMIN_01', 'SYSADMIN_02', 'SYSADMIN_03'],
    'User Access Management': ['UAM_01', 'UAM_02', 'UAM_03'],
    'Audit Log Access': ['AUDIT_VIEW_01', 'AUDIT_VIEW_02', 'AUDIT_VIEW_03'],
    'Modify Audit Logs': ['AUDIT_MODIFY_01', 'AUDIT_MODIFY_02', 'AUDIT_MODIFY_03'],
    'Generate Compliance Reports': ['COMPLIANCE_REPORT_01', 'COMPLIANCE_REPORT_02', 'COMPLIANCE_REPORT_03'],
    'Authorize Transactions': ['TXN_AUTH_01', 'TXN_AUTH_02', 'TXN_AUTH_03'],
    'View Sensitive Data': ['VIEW_SENSITIVE_01', 'VIEW_SENSITIVE_02', 'VIEW_SENSITIVE_03'],
    'Export Financial Data': ['EXPORT_FIN_01', 'EXPORT_FIN_02', 'EXPORT_FIN_03'],
    'Create Purchase Requisitions': ['PREQ_CREATE_01', 'PREQ_CREATE_02', 'PREQ_CREATE_03'],
    'Approve Purchase Requisitions': ['PREQ_APPROVE_01', 'PREQ_APPROVE_02', 'PREQ_APPROVE_03'],
    'Release Purchase Orders': ['PO_RELEASE_01', 'PO_RELEASE_02', 'PO_RELEASE_03']
}

MODULES = [
    'Procurement', 'Finance', 'IT Security', 'HR/Payroll',
    'Accounts Receivable', 'Accounts Payable', 'Inventory',
    'Supply Chain', 'Fixed Assets', 'Compliance', 'Audit'
]

RISK_RANKINGS = ['High', 'Medium', 'Low']

CONTROL_BUCKETS = [
    'Procurement Controls', 'Finance Controls', 'HR Controls',
    'IT Security Controls', 'Compliance Controls', 'Audit Controls',
    'Segregation of Duties', 'Approval Workflows', 'Access Controls'
]

RISK_DESCRIPTIONS_SOD = [
    'User can create and approve own transactions',
    'User can modify master data and authorize changes',
    'Segregation of duties violation in transaction cycle',
    'User can initiate and approve own requests',
    'Dual-role conflict in financial processes',
    'User can override system controls',
    'Unauthorized access to sensitive financial data',
    'User can process and reconcile own transactions'
]

RISK_DESCRIPTIONS_SA = [
    'Sensitive access – full system admin',
    'Sensitive access – payroll modification',
    'Sensitive access – system security controls',
    'Sensitive access – audit log modification',
    'Sensitive access – financial master data',
    'Sensitive access – compliance reports',
    'Sensitive access – user access management',
    'Sensitive access – password reset capability'
]

BUCKET_RECOMMENDATIONS = [
    'Ensure segregation of duties in transaction cycle. Implement dual approval.',
    'Enforce approval workflows and independent review for all transactions.',
    'Require manager sign-off on all critical operations.',
    'Implement monthly reconciliation and variance analysis.',
    'Restrict access based on job responsibilities.',
    'Maintain audit trail for all system changes.',
    'Conduct quarterly access review and recertification.',
    'Enforce strong authentication for sensitive operations.'
]

def generate_role_hierarchy(entitlement_privilege_map):
    """Generate role hierarchy with privileges based on entitlements."""

    # Define role templates with their typical entitlements
    role_templates = {
        'PROC_CLERK': ['Create Purchase Orders', 'Create Purchase Requisitions'],
        'PROC_MGR': ['Create Purchase Orders', 'Approve Purchase Orders', 'Create Vendor Master'],
        'PROC_DIR': ['Approve Purchase Orders', 'Release Purchase Orders', 'View Sensitive Data'],
        'FIN_ANALYST': ['Create Invoices', 'Post Journal Entries', 'Create Fixed Assets'],
        'FIN_MGR': ['Approve Invoices', 'Approve Journal Entries', 'Access Bank Accounts'],
        'FIN_DIR': ['Close General Ledger', 'Reconcile Bank Accounts', 'Authorize Transactions'],
        'FIN_CONTROLLER': ['System Administration', 'Export Financial Data', 'View Sensitive Data'],
        'HR_SPECIALIST': ['Create Employee Records', 'Process Payroll'],
        'HR_MGR': ['Modify Employee Records', 'Approve Payroll', 'Create Employee Records'],
        'HR_DIR': ['Authorize Transactions', 'Generate Compliance Reports'],
        'IT_ADMIN': ['System Administration', 'User Access Management'],
        'IT_SECURITY': ['User Access Management', 'Audit Log Access', 'System Administration'],
        'AUDITOR': ['Audit Log Access', 'Generate Compliance Reports', 'View Sensitive Data'],
        'COMPLIANCE_MGR': ['Generate Compliance Reports', 'View Sensitive Data', 'Audit Log Access'],
    }

    # Expand to 100 roles by creating variations
    expanded_roles = []
    base_roles = list(role_templates.keys())

    # Add base roles
    for role_code in base_roles:
        expanded_roles.append((role_code, role_code.replace('_', ' ').title()))

    # Create derived roles
    for i in range(len(base_roles), 100):
        base_role = random.choice(base_roles)
        dept = random.choice(['PROC', 'FIN', 'HR', 'IT', 'AUDIT'])
        level = random.choice(['L1', 'L2', 'L3', 'L4', 'L5'])
        role_code = f"{dept}_{level}_{i:03d}"
        role_name = f"{dept} {level} Role {i:03d}"
        expanded_roles.append((role_code, role_name))

    # Build hierarchy: assign entitlements to roles
    role_entitlements = {}

    # Map template roles to their entitlements
    for role_code, _ in expanded_roles:
        role_entitlements[role_code] = []

        # Check if it's a base role
        if role_code in role_templates:
            role_entitlements[role_code] = role_templates[role_code].copy()
        else:
            # Derived roles: assign random entitlements (some may be conflicting)
            # Higher chance of assigning conflicting combinations
            num_ents = random.randint(2, 4)
            role_entitlements[role_code] = random.sample(ENTITLEMENTS, num_ents)

    # Now build the actual role hierarchy data with privileges
    data = []
    for role_code, role_name in expanded_roles:
        entitlements = role_entitlements[role_code]

        for entitlement in entitlements:
            privileges = entitlement_privilege_map.get(entitlement, [])

            for priv_code in privileges:
                # Derive privilege display name from code
                priv_name = priv_code.replace('_', ' ').title()

                data.append({
                    'TOP_ROLE_CODE': role_code,
                    'TOP_ROLE_NAME': role_name,
                    'ROLE_CODE': None,
                    'ROLE_NAME': None,
                    'PRIVILEGE_CODE': priv_code,
                    'PRIVILEGE_NAME': priv_name
                })

    return data, expanded_roles, role_entitlements

def generate_user_roles(expanded_roles, role_entitlements, num_users=2000):
    """Generate user to role assignments, ensuring some users get conflicting roles."""
    data = []
    user_ids = [f"user.{i:04d}" for i in range(1, num_users + 1)]

    # Define conflicting role pairs that will trigger SOD violations
    conflicting_pairs = [
        ('PROC_CLERK', 'PROC_MGR'),  # Create and Approve PO
        ('FIN_ANALYST', 'FIN_MGR'),  # Create and Approve invoices
        ('HR_SPECIALIST', 'HR_MGR'),  # Create and Approve payroll
        ('IT_ADMIN', 'IT_SECURITY'),  # System admin and audit access
    ]

    for i, user_id in enumerate(user_ids):
        # 10% of users get conflicting role pairs
        if i % 10 == 0 and len(conflicting_pairs) > 0:
            pair = random.choice(conflicting_pairs)
            for role_code, role_name_template in expanded_roles:
                if role_code in pair:
                    for role_code2, role_name2 in expanded_roles:
                        if role_code2 == pair[1]:
                            data.append({
                                'User Name': user_id,
                                'Assigned Role Name': role_code,
                                'Assigned Role Display Name': f"{role_code} - {role_code2}"
                            })
                            data.append({
                                'User Name': user_id,
                                'Assigned Role Name': role_code2,
                                'Assigned Role Display Name': role_name2
                            })
                            break
                    break
        else:
            # Regular users: 1-3 non-conflicting roles
            num_roles = random.randint(1, 3)
            selected_roles = random.sample(expanded_roles, num_roles)

            for role_code, role_name in selected_roles:
                data.append({
                    'User Name': user_id,
                    'Assigned Role Name': role_code,
                    'Assigned Role Display Name': role_name
                })

    return data

def generate_ruleset(num_sod=50, num_sa=50):
    """Generate SOD and SA rulesets based on real entitlement conflicts."""

    # Define specific SOD conflicts (create/approve pairs, initiate/authorize pairs, etc.)
    sod_conflicts = [
        ('Create Purchase Orders', 'Approve Purchase Orders'),
        ('Create Purchase Orders', 'Release Purchase Orders'),
        ('Create Vendor Master', 'Approve Invoices'),
        ('Create Invoices', 'Approve Invoices'),
        ('Post Journal Entries', 'Approve Journal Entries'),
        ('Create Employee Records', 'Approve Payroll'),
        ('Process Payroll', 'Approve Payroll'),
        ('Access Bank Accounts', 'Reconcile Bank Accounts'),
        ('Create Fixed Assets', 'Depreciate Fixed Assets'),
        ('Create Accounts Receivable', 'Collect Accounts Receivable'),
    ]

    sod_data = []
    for i in range(1, num_sod + 1):
        if i <= len(sod_conflicts):
            lhs_ent, rhs_ent = sod_conflicts[i - 1]
        else:
            lhs_ent, rhs_ent = random.sample(ENTITLEMENTS, 2)

        sod_data.append({
            'Control Name': f'SOD-{i:03d}',
            'Risk Ranking': random.choice(RISK_RANKINGS),
            'LHS Entitlement': lhs_ent,
            'RHS Entitlement': rhs_ent,
            'Module(s)': random.choice(MODULES),
            'Risk Description': random.choice(RISK_DESCRIPTIONS_SOD),
            'Control Bucket': random.choice(CONTROL_BUCKETS)
        })

    # SA conflicts: sensitive access
    sa_conflicts = [
        'System Administration',
        'User Access Management',
        'Modify Audit Logs',
        'View Sensitive Data',
        'Export Financial Data',
        'Authorize Transactions',
    ]

    sa_data = []
    for i in range(1, num_sa + 1):
        if i <= len(sa_conflicts):
            entitlement = sa_conflicts[i - 1]
            side = 'LHS'
        else:
            entitlement = random.choice(ENTITLEMENTS)
            side = random.choice(['LHS', 'RHS'])

        sa_data.append({
            'Control Name': f'SA-{i:03d}',
            'Risk Ranking': random.choice(RISK_RANKINGS),
            'Entitlement': entitlement,
            'Side': side,
            'Module(s)': random.choice(MODULES),
            'Risk Description': random.choice(RISK_DESCRIPTIONS_SA)
        })

    return sod_data, sa_data

def generate_fp_database():
    """Generate FP database sheets."""

    # No action privileges (view-only, harmless)
    no_action_data = []
    reasons = [
        'View-only access – no transaction risk',
        'Read-only access – approved for auditors',
        'Display-only – no data modification',
        'Informational access – no system impact',
        'Query-only – no write capability'
    ]

    for ent_key in list(ENTITLEMENT_PRIVILEGE_MAP.keys())[:10]:  # First 10 entitlements
        privs = ENTITLEMENT_PRIVILEGE_MAP[ent_key]
        if privs:
            no_action_data.append({
                'PRIVILEGE_NAME': privs[0],  # Just mark the first privilege as view-only
                'False Positive Reason': random.choice(reasons)
            })

    # Work area privileges (department-based)
    work_area_data = []
    work_areas = ['WA_PROC', 'WA_FIN', 'WA_HR', 'WA_IT', 'WA_AUDIT']

    for ent_key, privs in list(ENTITLEMENT_PRIVILEGE_MAP.items())[:20]:
        if privs:
            work_area_data.append({
                'PRIVILEGE_NAME': privs[0],
                'WORK_AREA_PRIVILEGE_CODE': f"{random.choice(work_areas)}_{random.randint(1, 10):02d}"
            })

    return no_action_data, work_area_data

def write_xlsx(filepath, sheets_data):
    """Write data to XLSX file."""
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, data in sheets_data:
        ws = wb.create_sheet(sheet_name)

        if not data:
            continue

        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx).value = header

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx).value = row_data.get(header)

    wb.save(filepath)
    print(f"Generated: {filepath}")

def main():
    output_dir = Path(__file__).parent.parent / 'Data' / 'SODSAAnalysis'
    output_dir.mkdir(parents=True, exist_ok=True)

    print("Generating expanded SOD & SA Analysis data with real conflicts...")
    print("  - 50 SOD controls (with real entitlement conflicts)")
    print("  - 50 SA controls (with sensitive access rules)")
    print("  - 100 roles (with conflicting entitlements)")
    print("  - 2000 users (10% with conflicting role pairs)")

    # Generate data
    print("\n1. Building entitlement-to-privilege mapping...")
    print(f"   {len(ENTITLEMENT_PRIVILEGE_MAP)} entitlements mapped to privileges")

    print("2. Generating role hierarchy with conflicting entitlements...")
    hierarchy_data, expanded_roles, role_entitlements = generate_role_hierarchy(ENTITLEMENT_PRIVILEGE_MAP)
    print(f"   Created {len(expanded_roles)} roles with conflicting entitlements")

    print("3. Generating user-role assignments (10% with conflicts)...")
    user_roles_data = generate_user_roles(expanded_roles, role_entitlements, num_users=2000)
    print(f"   Created {len(user_roles_data)} user-role assignments")

    print("4. Generating SOD & SA rulesets from real conflicts...")
    sod_data, sa_data = generate_ruleset(50, 50)
    print(f"   Created {len(sod_data)} SOD controls + {len(sa_data)} SA controls")

    print("5. Generating FP database...")
    no_action_data, work_area_data = generate_fp_database()
    print(f"   Created {len(no_action_data)} no-action privileges + {len(work_area_data)} work-area privileges")

    # Write files
    print("\n6. Writing Excel files...")

    # Role hierarchy
    write_xlsx(
        output_dir / 'role_hierarchy.xlsx',
        [('Sheet1', hierarchy_data)]
    )

    # User roles
    write_xlsx(
        output_dir / 'user_roles.xlsx',
        [('Sheet1', user_roles_data)]
    )

    # Ruleset
    write_xlsx(
        output_dir / 'ruleset.xlsx',
        [
            ('SoD Ruleset', sod_data),
            ('SA Ruleset', sa_data),
            ('Entitlement to Privilege', _generate_entitlement_mapping(ENTITLEMENT_PRIVILEGE_MAP)),
            ('Bucket Details', _generate_bucket_details())
        ]
    )

    # FP database
    write_xlsx(
        output_dir / 'fp_database.xlsx',
        [
            ('No_action_Privileges', no_action_data),
            ('WorkArea_Privileges', work_area_data)
        ]
    )

    print("\nDone! All files generated with expanded data and real SOD conflicts.")

def _generate_entitlement_mapping(entitlement_privilege_map):
    """Convert entitlement-privilege map to data format."""
    data = []
    for entitlement, privileges in entitlement_privilege_map.items():
        for priv_code in privileges:
            data.append({
                'Entitlement Name': entitlement,
                'Privilege Code': priv_code
            })
    return data

def _generate_bucket_details():
    """Generate bucket details."""
    data = []
    for bucket in CONTROL_BUCKETS:
        data.append({
            'Bucket Name': bucket,
            'Risk': random.choice(['High', 'Medium', 'Low']),
            'EY Recommendations': random.choice(BUCKET_RECOMMENDATIONS)
        })
    return data

if __name__ == '__main__':
    main()
