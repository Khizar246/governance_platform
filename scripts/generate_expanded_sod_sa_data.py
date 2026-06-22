#!/usr/bin/env python3
"""Generate expanded SOD & SA Analysis datasets with larger volume for review."""

import random
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Seed for reproducibility
random.seed(42)

# Base data
MODULES = [
    'Procurement', 'Finance', 'IT Security', 'HR/Payroll',
    'Accounts Receivable', 'Accounts Payable', 'Inventory',
    'Supply Chain', 'Fixed Assets', 'Compliance', 'Audit'
]

RISK_RANKINGS = ['High', 'Medium', 'Low']

CONTROL_BUCKETS = [
    'Procurement Controls', 'Finance Controls', 'HR Controls',
    'IT Security Controls', 'Compliance Controls', 'Audit Controls',
    'Segregation of Duties', 'Approval Workflows', 'Access Controls'
]

PRIVILEGES_BASE = [
    ('CREATE', 'Create'), ('APPROVE', 'Approve'), ('MODIFY', 'Modify'),
    ('DELETE', 'Delete'), ('VIEW', 'View'), ('EXPORT', 'Export'),
    ('ADMIN', 'Administer'), ('RECONCILE', 'Reconcile'), ('AUTHORIZE', 'Authorize'),
    ('POST', 'Post'), ('REVERSE', 'Reverse'), ('RELEASE', 'Release'),
    ('REVIEW', 'Review'), ('APPROVE_FINAL', 'Final Approve'), ('CLOSE', 'Close')
]

ENTITLEMENTS = [
    'Purchase Orders', 'Vendor Master', 'Invoice Processing',
    'General Ledger', 'Journal Entry', 'Payment Processing',
    'Employee Records', 'Payroll Processing', 'System Administration',
    'User Access Management', 'Inventory Movement', 'Compliance Reports',
    'Accounts Receivable', 'Collections', 'Fixed Assets',
    'Bank Reconciliation', 'Intercompany Transactions', 'Budget Management',
    'Tax Processing', 'Audit Logs', 'System Monitoring'
]

RISK_DESCRIPTIONS_SOD = [
    'User can create and approve own transactions',
    'User can modify master data and authorize changes',
    'Segregation of duties violation in transaction cycle',
    'User can initiate and approve own requests',
    'Dual-role conflict in financial processes',
    'User can override system controls',
    'Unauthorized access to sensitive financial data',
    'User can process and reconcile own transactions'
]

RISK_DESCRIPTIONS_SA = [
    'Sensitive access – full system admin',
    'Sensitive access – payroll modification',
    'Sensitive access – system security controls',
    'Sensitive access – audit log modification',
    'Sensitive access – financial master data',
    'Sensitive access – compliance reports',
    'Sensitive access – user access management',
    'Sensitive access – password reset capability'
]

BUCKET_RECOMMENDATIONS = [
    'Ensure segregation of duties in transaction cycle. Implement dual approval.',
    'Enforce approval workflows and independent review for all transactions.',
    'Require manager sign-off on all critical operations.',
    'Implement monthly reconciliation and variance analysis.',
    'Restrict access based on job responsibilities.',
    'Maintain audit trail for all system changes.',
    'Conduct quarterly access review and recertification.',
    'Enforce strong authentication for sensitive operations.'
]

def generate_role_hierarchy(num_sod_controls, num_sa_controls):
    """Generate role hierarchy with privileges."""
    # Create base roles
    base_roles = [
        ('PROC_CLERK', 'Procurement Clerk'),
        ('PROC_MGR', 'Procurement Manager'),
        ('FIN_ANALYST', 'Financial Analyst'),
        ('FIN_MGR', 'Finance Manager'),
        ('FIN_DIR', 'Finance Director'),
        ('HR_SPECIALIST', 'HR Specialist'),
        ('HR_MGR', 'HR Manager'),
        ('IT_ADMIN', 'IT Administrator'),
        ('IT_SECURITY', 'IT Security Officer'),
        ('AUDITOR', 'Internal Auditor'),
        ('COMPLIANCE_MGR', 'Compliance Manager'),
        ('VENDOR_MGR', 'Vendor Manager'),
        ('AR_SPECIALIST', 'AR Specialist'),
        ('AP_SPECIALIST', 'AP Specialist'),
        ('INVENTORY_MGR', 'Inventory Manager'),
    ]

    # Expand to 100 roles
    expanded_roles = []
    expanded_roles.extend(base_roles)

    for i in range(len(base_roles), 100):
        dept = random.choice(['Procurement', 'Finance', 'HR', 'IT', 'Operations', 'Compliance'])
        level = random.choice(['Specialist', 'Analyst', 'Manager', 'Officer', 'Lead'])
        role_code = f"{dept[:3].upper()}_{level[:3].upper()}_{i:03d}"
        role_name = f"{dept} {level} {i:03d}"
        expanded_roles.append((role_code, role_name))

    # Create privilege list
    privileges = []
    for i in range(1, 200):  # 200 privileges
        module = random.choice(MODULES)
        action, action_name = random.choice(PRIVILEGES_BASE)
        entity = random.choice(['PO', 'Invoice', 'Journal', 'Employee', 'Vendor', 'GL', 'Payroll', 'Report'])
        priv_code = f"{entity}_{action}_{i:02d}"
        priv_name = f"{action_name} {entity} {i:02d}"
        privileges.append((priv_code, priv_name))

    # Build hierarchy: each role gets 3-8 privileges
    data = []
    for role_code, role_name in expanded_roles:
        top_role_code = role_code
        top_role_name = role_name

        num_privs = random.randint(3, 8)
        selected_privs = random.sample(privileges, num_privs)

        for priv_code, priv_name in selected_privs:
            data.append({
                'TOP_ROLE_CODE': top_role_code,
                'TOP_ROLE_NAME': top_role_name,
                'ROLE_CODE': None,
                'ROLE_NAME': None,
                'PRIVILEGE_CODE': priv_code,
                'PRIVILEGE_NAME': priv_name
            })

    return data, expanded_roles, privileges

def generate_user_roles(expanded_roles, num_users=2000):
    """Generate user to role assignments."""
    data = []
    user_ids = [f"user.{i:04d}" for i in range(1, num_users + 1)]

    for user_id in user_ids:
        # Each user gets 1-3 roles
        num_roles = random.randint(1, 3)
        selected_roles = random.sample(expanded_roles, num_roles)

        for role_code, role_name in selected_roles:
            data.append({
                'User Name': user_id,
                'Assigned Role Name': role_code,
                'Assigned Role Display Name': role_name
            })

    return data

def generate_ruleset(num_sod=50, num_sa=50):
    """Generate SOD and SA rulesets."""

    # Generate SOD controls
    sod_data = []
    for i in range(1, num_sod + 1):
        ent1, ent2 = random.sample(ENTITLEMENTS, 2)
        sod_data.append({
            'Control Name': f'SOD-{i:03d}',
            'Risk Ranking': random.choice(RISK_RANKINGS),
            'LHS Entitlement': ent1,
            'RHS Entitlement': ent2,
            'Module(s)': random.choice(MODULES),
            'Risk Description': random.choice(RISK_DESCRIPTIONS_SOD),
            'Control Bucket': random.choice(CONTROL_BUCKETS)
        })

    # Generate SA controls
    sa_data = []
    for i in range(1, num_sa + 1):
        sa_data.append({
            'Control Name': f'SA-{i:03d}',
            'Risk Ranking': random.choice(RISK_RANKINGS),
            'Entitlement': random.choice(ENTITLEMENTS),
            'Side': random.choice(['LHS', 'RHS']),
            'Module(s)': random.choice(MODULES),
            'Risk Description': random.choice(RISK_DESCRIPTIONS_SA)
        })

    # Generate Entitlement to Privilege mapping
    entitlement_mapping = []
    for entitlement in ENTITLEMENTS:
        num_privs = random.randint(2, 5)
        for _ in range(num_privs):
            entitlement_mapping.append({
                'Entitlement Name': entitlement,
                'Privilege Code': f"PRIV_{random.randint(1, 500):03d}"
            })

    # Generate Bucket Details
    bucket_details = []
    for bucket in CONTROL_BUCKETS:
        bucket_details.append({
            'Bucket Name': bucket,
            'Risk': random.choice(['High', 'Medium', 'Low']),
            'EY Recommendations': random.choice(BUCKET_RECOMMENDATIONS)
        })

    return sod_data, sa_data, entitlement_mapping, bucket_details

def generate_fp_database():
    """Generate FP database sheets."""

    # No action privileges
    no_action_data = []
    reasons = [
        'View-only access – no transaction risk',
        'Read-only access – approved for auditors',
        'Display-only – no data modification',
        'Informational access – no system impact',
        'Query-only – no write capability'
    ]

    for i in range(1, 31):
        no_action_data.append({
            'PRIVILEGE_NAME': f'VIEW_PRIV_{i:02d}',
            'False Positive Reason': random.choice(reasons)
        })

    # Work area privileges
    work_area_data = []
    for i in range(1, 51):
        work_area_data.append({
            'PRIVILEGE_NAME': f'PRIV_{i:03d}',
            'WORK_AREA_PRIVILEGE_CODE': f'WA_{random.randint(1, 100):02d}'
        })

    return no_action_data, work_area_data

def write_xlsx(filepath, sheets_data):
    """Write data to XLSX file."""
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for sheet_name, data in sheets_data:
        ws = wb.create_sheet(sheet_name)

        if not data:
            continue

        # Write headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header

        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = row_data.get(header)

    wb.save(filepath)
    print(f"Generated: {filepath}")

def main():
    output_dir = Path(__file__).parent.parent / 'Data' / 'SODSAAnalysis'
    output_dir.mkdir(parents=True, exist_ok=True)

    print("Generating expanded SOD & SA Analysis data...")
    print("  - 50 SOD controls")
    print("  - 50 SA controls")
    print("  - 100 roles")
    print("  - 2000 users")

    # Generate data
    print("\n1. Generating role hierarchy...")
    hierarchy_data, expanded_roles, privileges = generate_role_hierarchy(50, 50)

    print(f"   Created {len(expanded_roles)} roles with {len(privileges)} privileges")

    print("2. Generating user-role assignments...")
    user_roles_data = generate_user_roles(expanded_roles, num_users=2000)
    print(f"   Created {len(user_roles_data)} user-role assignments")

    print("3. Generating rulesets...")
    sod_data, sa_data, entitlement_data, bucket_data = generate_ruleset(50, 50)
    print(f"   Created {len(sod_data)} SOD controls + {len(sa_data)} SA controls")

    print("4. Generating FP database...")
    no_action_data, work_area_data = generate_fp_database()
    print(f"   Created {len(no_action_data)} no-action privileges + {len(work_area_data)} work-area privileges")

    # Write files
    print("\n5. Writing Excel files...")

    # Role hierarchy
    write_xlsx(
        output_dir / 'role_hierarchy.xlsx',
        [('Sheet1', hierarchy_data)]
    )

    # User roles
    write_xlsx(
        output_dir / 'user_roles.xlsx',
        [('Sheet1', user_roles_data)]
    )

    # Ruleset
    write_xlsx(
        output_dir / 'ruleset.xlsx',
        [
            ('SoD Ruleset', sod_data),
            ('SA Ruleset', sa_data),
            ('Entitlement to Privilege', entitlement_data),
            ('Bucket Details', bucket_data)
        ]
    )

    # FP database
    write_xlsx(
        output_dir / 'fp_database.xlsx',
        [
            ('No_action_Privileges', no_action_data),
            ('WorkArea_Privileges', work_area_data)
        ]
    )

    print("\nDone! All files generated with expanded data.")

if __name__ == '__main__':
    main()
